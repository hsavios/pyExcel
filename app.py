import os
import customtkinter as ctk
from openpyxl import Workbook, load_workbook
import pyperclip
import time


class Application(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Cadastro de dados")
        self.geometry("300x250")

        self.date_label = ctk.CTkLabel(self, text="Data:")
        self.date_label.pack()

        self.date_entry = ctk.CTkEntry(self)
        self.date_entry.pack()

        self.plate_label = ctk.CTkLabel(self, text="Placa:")
        self.plate_label.pack()

        self.plate_entry = ctk.CTkEntry(self)
        self.plate_entry.pack()

        self.state_label = ctk.CTkLabel(self, text="Estado:")
        self.state_label.pack()

        self.state_entry = ctk.CTkEntry(self)
        self.state_entry.pack()

        self.weight_label = ctk.CTkLabel(self, text="Peso:")
        self.weight_label.pack()

        self.weight_entry = ctk.CTkEntry(self)
        self.weight_entry.pack()

        self.save_button = ctk.CTkButton(
            self, text="Salvar", command=self.save_data)
        self.save_button.pack()

        self.total_weight_label = ctk.CTkLabel(self, text="Total de peso: 0")
        self.total_weight_label.pack()

        self.file_name = "dados.xlsx"
        if os.path.isfile(self.file_name):
            self.wb = load_workbook(self.file_name)
            self.ws = self.wb.active
            self.update_total_weight()
        else:
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.append(["Data", "Placa", "Estado", "Peso"])
            self.total_weight = 0

    def update_total_weight(self):
        total_weight = sum([float(cell.value) for cell in self.ws['D'][1:]])
        self.total_weight_label.configure(text=f"Total de peso: {total_weight}")

    def save_data(self):
        date = self.date_entry.get()
        plate = self.plate_entry.get()
        state = self.state_entry.get()
        weight = self.weight_entry.get()

        self.ws.append([date, plate, state, weight])
        self.wb.save(self.file_name)

        pyperclip.copy(f"{plate}")
        time.sleep(3)
        pyperclip.copy(f"{state}")
        time.sleep(3)
        pyperclip.copy(f"{weight}")

        self.update_total_weight()


app = Application()
app.mainloop()
